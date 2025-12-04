# routes_export.py
from pathlib import Path
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, case

from database import get_db
from security import require_roles_dep
from logging_config import logger
from models import Movimiento

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional


# ============================
#   TEMPLATES
# ============================

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


# ============================
#      ROUTER EXPORT
# ============================

router = APIRouter(
    prefix="",          # igual que en routes_alerts
    tags=["export"],
)


# =====================================================
# VISTA HTML: HOME DE EXPORTACIÓN
# =====================================================

@router.get("/exportar", response_class=HTMLResponse)
async def export_home(
    request: Request,
    user: dict = Depends(require_roles_dep("admin", "superadmin")),
):
    """
    Pantalla de exportación para el negocio.
    Muestra opciones para descargar stock actual y movimientos.
    """
    return templates.TemplateResponse(
        "exportar.html",
        {
            "request": request,
            "user": user,
        },
    )


# =====================================================
# Utilidad genérica para construir un Excel en memoria
# =====================================================

def build_excel(headers: list[str], rows: list[tuple], title: str = "Reporte"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limita a 31 caracteres

    # Encabezados
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Filas
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-ajustar ancho de columnas (simple)
    for col_idx, _ in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 18

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# =====================================================
# EXPORT: STOCK ACTUAL
# =====================================================

@router.get("/exportar/stock")
async def export_stock_actual(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_roles_dep("admin", "superadmin")),
):
    """
    Exporta el stock actual por producto y slot (código_full) a Excel.

    - Se basa en la tabla movimientos.
    - Asume que Movimiento.tipo indica si la cantidad suma o resta:
        - tipo = 'salida'  -> cantidad negativa
        - otros tipos      -> cantidad positiva (entrada, ajuste+, etc.)
    - Filtra por negocio_id del usuario.
    """

    negocio_id = user.get("negocio_id")

    # Cantidad neta con signo según tipo
    cantidad_neta = case(
        (Movimiento.tipo == "salida", -Movimiento.cantidad),
        else_=Movimiento.cantidad,
    )

    query = (
        db.query(
            Movimiento.producto.label("producto"),
            Movimiento.zona.label("slot_codigo_full"),
            func.sum(cantidad_neta).label("stock"),
        )
        .filter(Movimiento.negocio_id == negocio_id)
        .group_by(
            Movimiento.producto,
            Movimiento.zona,
        )
        .having(func.sum(cantidad_neta) != 0)
        .order_by(Movimiento.producto, Movimiento.zona)
    )

    resultados = query.all()

    headers = ["Producto", "Slot (código full)", "Stock actual"]
    rows = [
        (
            r.producto,
            r.slot_codigo_full,
            int(r.stock) if r.stock is not None else 0,
        )
        for r in resultados
    ]

    if not rows:
        logger.info(f"[EXPORT_STOCK] Sin resultados para negocio_id={negocio_id}")

    stream = build_excel(headers, rows, title="Stock actual")

    filename = f"stock_actual_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


# =====================================================
# EXPORT: MOVIMIENTOS
# =====================================================

@router.get("/exportar/movimientos")
async def export_movimientos(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_roles_dep("admin", "superadmin")),
    start_date: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
):
    """
    Exporta movimientos de stock a Excel en un rango de fechas.

    - Usa directamente los campos de Movimiento:
        fecha, tipo, producto, cantidad, zona (código_full), usuario,
        fecha_vencimiento, motivo_salida.
    - Filtra por negocio_id del usuario.
    - Si no se envía rango de fechas (o vienen vacías), exporta todos los movimientos.
    """

    negocio_id = user.get("negocio_id")

    # Parseo suave de fechas (admite "", None → None)
    start_dt: datetime | None = None
    end_dt: datetime | None = None

    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            logger.warning(f"[EXPORT_MOV] start_date inválida: {start_date}")
            start_dt = None

    if end_date:
        try:
            # usamos max.time() para incluir todo el día
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            end_dt = datetime.combine(end_dt.date(), datetime.max.time())
        except ValueError:
            logger.warning(f"[EXPORT_MOV] end_date inválida: {end_date}")
            end_dt = None

    q = (
        db.query(
            Movimiento.fecha,
            Movimiento.tipo,
            Movimiento.cantidad,
            Movimiento.producto,
            Movimiento.zona,
            Movimiento.usuario,
            Movimiento.fecha_vencimiento,
            Movimiento.motivo_salida,
        )
        .filter(Movimiento.negocio_id == negocio_id)
    )

    if start_dt:
        q = q.filter(Movimiento.fecha >= start_dt)
    if end_dt:
        q = q.filter(Movimiento.fecha <= end_dt)

    q = q.order_by(Movimiento.fecha.desc())
    resultados = q.all()

    headers = [
        "Fecha",
        "Tipo",
        "Cantidad",
        "Cantidad neta",           # con signo según tipo
        "Producto",
        "Slot (código full)",
        "Usuario",
        "Fecha vencimiento",
        "Motivo salida",
    ]

    rows: list[tuple] = []
    for r in resultados:
        # Cantidad neta por movimiento (misma lógica que en stock)
        if r.tipo == "salida":
            cantidad_neta = -(r.cantidad or 0)
        else:
            cantidad_neta = r.cantidad or 0

        rows.append(
            (
                r.fecha.isoformat() if r.fecha else None,
                r.tipo,
                int(r.cantidad) if r.cantidad is not None else 0,
                int(cantidad_neta),
                r.producto,
                r.zona,  # código_full
                r.usuario,
                r.fecha_vencimiento.isoformat() if r.fecha_vencimiento else None,
                r.motivo_salida,
            )
        )

    stream = build_excel(headers, rows, title="Movimientos")

    filename = f"movimientos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )
